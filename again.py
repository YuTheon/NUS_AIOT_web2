import os
import glob
from sklearn.preprocessing import MinMaxScaler
import numpy as np

scaler = MinMaxScaler(feature_range=(-1, 1))

def process2(file_name):
    # 获取文件夹中的所有文件
    s, e = -1, -1
    arrs = list()
    with open(file_name, 'r') as f:
        cons = f.read()
        arr = cons.split('\n')
        nums = len(arr)
        for i in range(nums):
            if arr[i] == 'S':
                s = i
            if arr[i] == 'E':
                e = i
        if e > 0:
            arr.pop(e)
        if s > 0:
            arr.pop(s)
    for i in arr:
        arrs.append(round(int(i), 3))
    arrs = np.array(arrs)
    
    arrs = scaler.fit_transform(arrs.reshape(-1, 1))
    return arrs, s, e

def process(folder):

    # 文件夹路径
    folder_path = folder

    # 获取文件夹中的所有文件
    files_ = glob.glob(os.path.join(folder_path, "*"))
    files = list()
    for file in files_:
        if file.endswith("txt"):
            files.append(file)
    print(f'files {len(files)}')

    # 遍历所有文件
    for file in files:
        # 判断文件是否是普通文件
        if os.path.isfile(file):
            # 获取文件名（不包含扩展名）
            name = os.path.splitext(os.path.basename(file))[0]
            import re


            file_name = name + ".txt"
            data_name = name + ".xlsx"
            file_path = os.path.join(folder_path, file_name)
            data_path = os.path.join(folder_path, data_name)
            print(file_path)
            print(data_path)
            with open(file_path, "r") as f:
                # 读取文件中的所有内容
                content = f.read()
                # 打印读取到的内容
            print(content)
            original_list = [s.strip() for s in content.split("\n")]
            print(original_list)
            # 遍历每个子串，并判断是否需要将空行填上 ZorY
            processed_list = []  # 处理后的列表，初始值为原始列表的第一个元素
            for i in range(0, len(original_list), 1):
                if i + 1 < len(original_list) and (
                        original_list[i].lstrip("-").isdigit() and original_list[i + 1].lstrip("-").isdigit()):
                    # 如果当前元素和下一个元素都是数字，则在它们之间插入“Y”或“Z”
                    if original_list[i - 1] == 'Y':
                        processed_list += [original_list[i], 'Z']
                    else:
                        processed_list += [original_list[i], 'Y']
                else:
                    # 如果当前元素和下一个元素不都是数字，则直接将它们添加到处理后的列表中
                    processed_list.append(original_list[i])

            print(processed_list)

            # 处理掉了一整个Y20或Z20的情况:删除掉与之配对的Z10或Y19
            processed_list2 = []  # 初始化处理后的列表
            i = 0
            while i < len(processed_list):
                if processed_list[i] == 'Y':
                    # 如果当前元素是“Y”，则判断下一个元素是否也是“Y”
                    if i + 2 < len(processed_list) and processed_list[i + 2] == 'Y':
                        # 如果下一个元素也是“Y”，则跳过当前元素和下一个元素，处理下一个元素
                        i += 2
                    else:
                        # 如果下一个元素不是“Y”，则将当前元素和下一个元素添加到处理后的列表中
                        processed_list2 += [processed_list[i], processed_list[i + 1]]
                        i += 2
                elif processed_list[i] == 'Z':
                    # 如果当前元素是“Z”，则判断下一个元素是否也是“Z”
                    if i + 2 < len(processed_list) and processed_list[i + 2] == 'Z':
                        # 如果下一个元素也是“Z”，则跳过当前元素和下一个元素，处理下一个元素
                        i += 2
                    else:
                        # 如果下一个元素不是“Z”，则将当前元素和下一个元素添加到处理后的列表中
                        processed_list2 += [processed_list[i], processed_list[i + 1]]
                        i += 2
                        # 如果下一个元素是“Y”，则将其添加到处理后的列表中
                else:
                    # 如果当前元素不是“Y”或“Z”，则将当前元素添加到处理后的列表中
                    processed_list2.append(processed_list[i])
                    i += 1

            print(processed_list2)

            # 使用 join() 方法将列表中的元素拼接成一个新的字符串
            result_str = "".join(processed_list2)

            # 打印拼接后的字符串
            print("content")
            print(result_str)

            # 假设传给 Python 的数据格式为字符串 data
            data = result_str
            # 切割头部
            first_y_pos = -1

            # 遍历字符串中的每个字符，查找第一个为 "y" 的字符
            for i in range(len(data)):
                if data[i] == "Y":
                    first_y_pos = i
                    break

            # 如果找到了第一个 "y"，则删除该字符及其之前的内容
            if first_y_pos >= 0 and data[first_y_pos - 1] != 'S':
                data = data[first_y_pos:]

            if first_y_pos >= 0 and data[first_y_pos - 1] == 'S':
                data = data[first_y_pos - 1:]
            # 切割尾部
            last_y_pos = -1
            last_z_pos = -1

            # 遍历字符串中的每个字符，查找最后一个 "y" 和 "z" 的位置
            for i in range(len(data)):
                if data[i] == "Y":
                    last_y_pos = i
                elif data[i] == "Z":
                    last_z_pos = i

            # 如果最后一个 "y" 的位置在最后一个 "z" 的位置之前，则截取 "y" 之前的内容
            if last_y_pos > last_z_pos or last_z_pos == len(data) - 1:
                data = data[:last_y_pos]

            print("已截取")
            print(data)

            start_num = 1
            end_num = 1

            i = 0
            flag = False

            # 进入 while 循环，读取字符串中的每个字符
            while i < len(data):
                # 如果当前字符是 "y"，则将 m 和 n 均加 1
                if data[i] == "Y":
                    if not flag:
                        start_num += 1
                    end_num += 1
                # 如果当前字符是 "s"，则停止增加 m
                elif data[i] == "S":
                    flag = True
                # 如果当前字符是 "e"，则停止修改 n 并跳出循环
                elif data[i] == "E":
                    break
                # 如果当前字符是其他字符，则不做处理
                i += 1

            print("start")
            print(start_num)
            print("end")
            print(end_num)

            data = data.replace("S", "").replace("E", "")
            print("已删除SE")
            print(data)
            data = list(data)
            ff = 1
            # 遍历每个子串，并判断是否需要在 Y 和 Z 中间填入 0
            for i in range(len(data)):
                if data[i] == "Z" and i + 1 < len(data) and data[i + 1] == "Y":
                    data.insert(i + 1, "0")
                    ff = 0
            print(ff)
            # 使用 join() 方法将列表中的元素拼接成一个新的字
            data = "".join(data)

            print("补全空缺")
            print(data)
            # 将字符串按字母 "y" 和 "z" 分割成列表
            data_list = data.split("Y")[1:]
            print(data_list)
            # 初始化两个空列表，用于存储 y 和 z 坐标
            y_coords = []
            z_coords = []
            # 处理只剩负号的情况
            # 遍历列表中的每个元素，提取 y 和 z 坐标，并分别存储到对应的列表中
            for i in range(len(data_list)):
                # 处理缺失到只剩负号的异常值
                data_list[i] = re.sub(r"Z-$", "Z-999", data_list[i])

            print("补负号后")
            print(data_list)
            for i in range(len(data_list)):
                # 先将逗号替换为空格，然后再将字符串分割成多个子字符串
                y_coord, z_coord = data_list[i].replace(",", "").split("Z")
                # 将子字符串转换为整数
                y_coords.append(int(y_coord))
                z_coords.append(int(z_coord))

            # 将 y 和 z 坐标列表合并成一个两列的矩阵
            coords_matrix = list(zip(y_coords, z_coords))

            # 打印结果
            print(coords_matrix)

            # 处理异常数据之处理-999
            z_sum = 0
            count = 0
            for coord in coords_matrix:
                if coord[1] != -999:
                    z_sum += coord[1]
                    count += 1
            z_mean = z_sum / count

            # 将所有的 -999 改写为平均值
            new_coords = []
            for coord in coords_matrix:
                if coord[1] == -999:
                    new_coords.append((coord[0], z_mean))
                else:
                    new_coords.append(coord)

            print(new_coords)

            # 定义一个滑动窗口大小
            window_size = 10

            # 初始化一个列表来保存每个窗口的平均值和标准差
            averages = []
            stds = []

            # 检测和修正异常值
            answer = []
            for i in range(len(new_coords)):
                # 将数据添加到滑动窗口中
                if i % window_size == 0:
                    # 计算当前窗口的平均值和标准差，并添加到列表中
                    window = new_coords[i:i + window_size]
                    y_values = [coord[0] for coord in window]
                    y_mean = sum(y_values) / len(y_values)
                    y_std = (sum([(y - y_mean) ** 2 for y in y_values]) / len(y_values)) ** 0.5
                    averages.append(y_mean)
                    stds.append(y_std)

                # 检查并修正异常值
                if abs(new_coords[i][0] - averages[-1]) > stds[-1]:
                    # 如果 y 坐标是异常值，则将其替换为平均值
                    new_y = averages[-1]
                else:
                    new_y = new_coords[i][0]

                # 将修正后的数据添加到答案列表中
                answer.append((new_y, new_coords[i][1]))

            print(answer)

            import pandas as pd

            # 将 answer 转换为 DataFrame 对象
            df = pd.DataFrame(answer, columns=['Y', 'Z'])

            # 将 DataFrame 导出到 Excel 文件中
            df.to_excel(data_path, index=False)

            # 将S、E导入excel
            from openpyxl import Workbook, load_workbook

            # 打开 Excel 文件
            wb = load_workbook(data_path)

            # 创建或选择一个名为 Sheet1 的工作表
            ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.create_sheet('Sheet1')

            # 将两个值分别导入到第三列和第四列的第一个位置

            ws.cell(row=2, column=3, value=start_num)
            ws.cell(row=1, column=3, value='S')
            ws.cell(row=2, column=4, value=end_num)
            ws.cell(row=1, column=4, value='E')

            # 保存 Excel 文件
            wb.save(data_path)

if __name__ == "__main__":
    process("./datadata/")